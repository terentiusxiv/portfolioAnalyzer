#!/usr/bin/env python3
"""
Lightweight Portfolio Analyzer
Supports Saxo Bank / Mandatum Trader PDF reports, Nordnet PDF reports,
Nordea XLSX exports, and a universal CSV format.
Enriches with Yahoo Finance for live prices, sector, geography, and correlation.

Usage:
  python portfolio_analyzer.py Portfolio_report.pdf
  python portfolio_analyzer.py Salkkuraportti.pdf
  python portfolio_analyzer.py portfolio.csv
  python portfolio_analyzer.py                       (auto-detects Portfolio_*.pdf/csv)
"""

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from abc import ABC, abstractmethod
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import yfinance as yf
from matplotlib.patches import Patch

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

ISIN_TICKER_MAP_FILE = "isin_ticker_map.json"
CASH_FLOWS_FILE = "cash_flows.json"
CHARTS_FILE = "portfolio_charts.png"
CORRELATION_CHARTS_FILE = "portfolio_correlation.png"
EXPANDED_CORRELATION_FILE = "portfolio_correlation_expanded.png"

DEFAULT_ISIN_MAP = {
    "US0079031078": "AMD",
    "FI4000512496": "CANATU.HE",
    "FI0009005870": "KCR.HE",
    "US5128073062": "LRCX",
    "FI4000552526": "MANTA.HE",
    "FI4000198031": "QTCOM.HE",
    "FI0009010912": "REV1V.HE",
    "US92532F1003": "VRTX",
}

ISIN_COUNTRY_PREFIXES = {"US", "FI", "DE", "GB", "FR", "NL", "SE", "DK", "NO", "IE", "CH", "LU", "BE", "JP", "CA"}

COLORS = ["#2563eb", "#16a34a", "#dc2626", "#9333ea", "#ea580c",
          "#0891b2", "#ca8a04", "#e11d48", "#4f46e5", "#059669"]


def parse_num(s):
    return float(s.strip().replace("\u00a0", "").replace(" ", "").replace("%", "").replace(",", "."))


class PortfolioParser(ABC):
    @abstractmethod
    def can_parse(self, path: str) -> bool:
        """Return True if this parser recognises the file."""

    @abstractmethod
    def parse(self, path: str) -> dict:
        """Return {"positions": [...], "cash_eur": float, "account_value_eur": float}."""


class SaxoParser(PortfolioParser):
    """Parses Saxo Bank / Mandatum Trader PDF portfolio reports."""

    def can_parse(self, path: str) -> bool:
        if Path(path).suffix.lower() != ".pdf":
            return False
        try:
            with pdfplumber.open(path) as pdf:
                text = (pdf.pages[0].extract_text() or "").lower()
            return "saxo" in text or "mandatum" in text
        except Exception:
            return True  # Accept any PDF as a fallback

    def parse(self, path: str) -> dict:
        if pdfplumber is None:
            print("pdfplumber not installed. Run: pip install pdfplumber")
            sys.exit(1)

        print(f"Parsing {path}...")

        with pdfplumber.open(path) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

        lines = full_text.split("\n")

        # Build line-level ISIN index
        isin_line_map = {}
        for i, line in enumerate(lines):
            for m in re.finditer(r"([A-Z]{2}[A-Z0-9]{10})", line):
                code = m.group(1)
                if code[:2] in ISIN_COUNTRY_PREFIXES:
                    isin_line_map[i] = code

        # Find data lines: start with EUR/USD, then integer quantity, then numbers
        data_lines = []
        for i, line in enumerate(lines):
            if re.match(r"^(EUR|USD)\s+\d+\s+[\d,]+\s+[\d,]+\s+[\d,]+", line):
                data_lines.append((i, line))

        # Match each data line to nearest ISIN within 3 lines
        positions = []
        used_isins = set()

        for dl_idx, dl_text in data_lines:
            best_isin, best_dist = None, 999
            for il_idx, isin in isin_line_map.items():
                dist = abs(il_idx - dl_idx)
                if dist < best_dist and dist <= 3 and isin not in used_isins:
                    best_dist = dist
                    best_isin = isin

            if not best_isin:
                continue
            used_isins.add(best_isin)

            parts = dl_text.split()
            currency = parts[0]
            quantity = int(parts[1])
            nums = re.findall(r"(-?[\d\s\u00a0]+,[\d]+)", dl_text)

            if len(nums) < 6:
                continue

            open_price = parse_num(nums[1])
            current_price = parse_num(nums[2])
            pnl = parse_num(nums[4])
            market_val = parse_num(nums[5])

            # Instrument name from line above
            name_line = lines[dl_idx - 1] if dl_idx > 0 else ""
            name = re.sub(r"\(ISIN:.*", "", name_line).strip()

            positions.append({
                "instrument": name, "isin": best_isin, "currency": currency,
                "quantity": quantity, "open_price": open_price, "current_price": current_price,
                "pnl_eur": pnl, "market_value_eur": market_val,
            })

        # Cash
        cash_eur = 0.0
        cash_match = re.search(r"Allaccounts\s+EUR\s+([\d\s\u00a0]+,\d{2})", full_text)
        if not cash_match:
            cash_match = re.search(r"All accounts\s+EUR\s+([\d\s\u00a0]+,\d{2})", full_text)
        if cash_match:
            cash_eur = parse_num(cash_match.group(1))

        equity_total = sum(p["market_value_eur"] for p in positions)
        account_value = equity_total + cash_eur

        print(f"   Extracted {len(positions)} stock positions")
        print(f"   Cash: EUR {cash_eur:,.2f}")
        print(f"   Account value: EUR {account_value:,.2f}\n")

        return {"positions": positions, "cash_eur": cash_eur, "account_value_eur": account_value}


class CsvParser(PortfolioParser):
    """Parses a universal CSV portfolio export.

    Required columns:
      instrument, isin, currency, quantity, open_price,
      current_price, pnl_eur, market_value_eur

    Cash balance: add a row with isin=CASH (or instrument=CASH) and
    set market_value_eur to the cash amount; all other fields can be 0.
    """

    REQUIRED_COLS = {"instrument", "isin", "currency", "quantity",
                     "open_price", "current_price", "pnl_eur", "market_value_eur"}

    def can_parse(self, path: str) -> bool:
        return Path(path).suffix.lower() == ".csv"

    def parse(self, path: str) -> dict:
        print(f"Parsing {path}...")
        df = pd.read_csv(path)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        missing = self.REQUIRED_COLS - set(df.columns)
        if missing:
            raise ValueError(f"CSV is missing required columns: {', '.join(sorted(missing))}")

        positions = []
        cash_eur = 0.0

        for _, row in df.iterrows():
            isin = str(row["isin"]).strip()
            instrument = str(row["instrument"]).strip()

            if isin.upper() == "CASH" or instrument.upper() == "CASH":
                cash_eur += float(row["market_value_eur"])
                continue

            positions.append({
                "instrument": instrument,
                "isin": isin,
                "currency": str(row["currency"]).strip(),
                "quantity": int(row["quantity"]),
                "open_price": float(row["open_price"]),
                "current_price": float(row["current_price"]),
                "pnl_eur": float(row["pnl_eur"]),
                "market_value_eur": float(row["market_value_eur"]),
            })

        equity_total = sum(p["market_value_eur"] for p in positions)
        account_value = equity_total + cash_eur

        print(f"   Loaded {len(positions)} positions from CSV")
        print(f"   Cash: EUR {cash_eur:,.2f}")
        print(f"   Account value: EUR {account_value:,.2f}\n")

        return {"positions": positions, "cash_eur": cash_eur, "account_value_eur": account_value}


class NordeaXlsxParser(PortfolioParser):
    """Parses Nordea portfolio exports (Holdings sheet from Omistukset.xlsx)."""

    def can_parse(self, path: str) -> bool:
        if Path(path).suffix.lower() != ".xlsx":
            return False
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Holdings" not in wb.sheetnames:
                return False
            ws = wb["Holdings"]
            rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            return bool(rows and rows[0][0] == "Type" and rows[0][1] == "AccountKey")
        except Exception:
            return False

    def parse(self, path: str) -> dict:
        if openpyxl is None:
            print("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)

        print(f"Parsing {path}...")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Holdings"]
        rows = list(ws.iter_rows(values_only=True))

        # Row 0 = timestamp, Row 1 = headers
        headers = {name: idx for idx, name in enumerate(rows[1]) if name is not None}

        def col(row, name):
            return row[headers[name]]

        positions, cash_eur = [], 0.0
        for row in rows[2:]:
            row_type = row[0]
            if row_type == "Custody":
                isin = col(row, "ISIN")
                if not isin:
                    continue
                positions.append({
                    "instrument": col(row, "NAME") or "",
                    "isin": str(isin).strip(),
                    "currency": col(row, "CURRENCY") or "EUR",
                    "quantity": int(col(row, "HOLDINGS") or 0),
                    "open_price": float(col(row, "Average purchase price") or 0),
                    "current_price": float(col(row, "PRICE") or 0),
                    "pnl_eur": float(col(row, "Value change on account level") or 0),
                    "market_value_eur": float(col(row, "Value on account level") or 0),
                })
            elif row_type == "CashAccount":
                cash_eur += float(col(row, "Value on account level") or 0)

        equity_total = sum(p["market_value_eur"] for p in positions)
        account_value = equity_total + cash_eur
        print(f"   Extracted {len(positions)} positions")
        print(f"   Cash: EUR {cash_eur:,.2f}")
        print(f"   Account value: EUR {account_value:,.2f}\n")
        return {"positions": positions, "cash_eur": cash_eur, "account_value_eur": account_value}


class NordnetPdfParser(PortfolioParser):
    """Parses Nordnet portfolio PDF reports (Salkkuraportti / Portfolio report).

    Columns in the holdings section:
      Name  MarketPrice+CCY  CostPrice  Qty  MarketValueEUR  UnrealizedResult  Weight%
    """

    # Currencies Nordnet reports are likely to contain
    _CURRENCIES = r"EUR|USD|DKK|NOK|SEK|GBP|CHF"

    # One holding row – parsed left-to-right, anchored at both ends.
    # The name is non-greedy so it stops at the first digits+currency token.
    _ROW_RE = re.compile(
        r"^(.+?)\s+"                                         # 1: instrument name
        r"([\d,]+)(" + r"EUR|USD|DKK|NOK|SEK|GBP|CHF" + r")\s+"  # 2: market price, 3: currency
        r"([\d,]+)\s+"                                       # 4: cost / avg buy price
        r"([\d,]+)\s+"                                       # 5: quantity (may be fractional)
        r"([\d ]+,\d{2})\s+"                                 # 6: market value in EUR
        r"([+\-\u2212][\d ,]+,\d{2})\s+"                    # 7: unrealised result
        r"([\d,]+)%$"                                        # 8: portfolio weight
    )

    @staticmethod
    def _pn(s: str) -> float:
        """Parse a Finnish-formatted number (comma decimal, space thousands, U+2212 minus)."""
        return float(
            s.strip()
            .replace("\u2212", "-")   # Finnish minus → ASCII hyphen
            .replace("\u00a0", "")    # non-breaking space
            .replace(" ", "")
            .replace(",", ".")
        )

    def can_parse(self, path: str) -> bool:
        if Path(path).suffix.lower() != ".pdf":
            return False
        if pdfplumber is None:
            return False
        try:
            with pdfplumber.open(path) as pdf:
                text = (pdf.pages[0].extract_text() or "").lower()
            return "nordnet" in text
        except Exception:
            return False

    def parse(self, path: str) -> dict:
        if pdfplumber is None:
            print("pdfplumber not installed. Run: pip install pdfplumber")
            sys.exit(1)

        print(f"Parsing {path}...")

        with pdfplumber.open(path) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

        positions = []
        for line in full_text.splitlines():
            m = self._ROW_RE.match(line.strip())
            if not m:
                continue
            name, mkt_price, ccy, cost_price, qty, mkt_val, result, _ = m.groups()
            positions.append({
                "instrument": name.strip(),
                "isin": "",           # not present in Nordnet PDF export
                "currency": ccy,
                "quantity": self._pn(qty),
                "open_price": self._pn(cost_price),
                "current_price": self._pn(mkt_price),
                "pnl_eur": self._pn(result),
                "market_value_eur": self._pn(mkt_val),
            })

        # Cash: "Likvidit varat  142,66  0,47%"
        cash_eur = 0.0
        cash_m = re.search(r"Likvidit varat\s+([\d ,]+,\d{2})\s+[\d,]+%", full_text)
        if cash_m:
            cash_eur = self._pn(cash_m.group(1))

        # Account total: "Yhteensä  30 639,69  +8 815,30  100,00%"
        total_m = re.search(r"Yhteens[äa]\s+([\d ,]+,\d{2})\s+[+\-\u2212]", full_text)
        if total_m:
            account_value = self._pn(total_m.group(1))
        else:
            account_value = sum(p["market_value_eur"] for p in positions) + cash_eur

        print(f"   Extracted {len(positions)} positions")
        print(f"   Cash: EUR {cash_eur:,.2f}")
        print(f"   Account value: EUR {account_value:,.2f}\n")

        return {"positions": positions, "cash_eur": cash_eur, "account_value_eur": account_value}


PARSERS: list[PortfolioParser] = [NordnetPdfParser(), SaxoParser(), NordeaXlsxParser(), CsvParser()]


def load_portfolio(path: str) -> dict:
    for parser in PARSERS:
        if parser.can_parse(path):
            return parser.parse(path)
    raise ValueError(
        f"No parser recognised '{path}'. "
        f"Supported formats: Nordnet PDF (Salkkuraportti), "
        f"Saxo Bank / Mandatum Trader PDF, Nordea XLSX (Omistukset.xlsx), CSV."
    )


def load_isin_map():
    if Path(ISIN_TICKER_MAP_FILE).exists():
        with open(ISIN_TICKER_MAP_FILE) as f:
            return json.load(f)
    return dict(DEFAULT_ISIN_MAP)


def save_isin_map(mapping):
    with open(ISIN_TICKER_MAP_FILE, "w") as f:
        json.dump(mapping, f, indent=2)


_ISIN_PREFIX_TO_EXCHCODES = {
    "US": {"UN", "UQ", "UA", "UR", "US"},
    "FI": {"HE"},
    "NL": {"AS"},
    "DE": {"GY", "DE"},
    "DK": {"DC", "CO"},
    "SE": {"SS", "ST"},
    "NO": {"OL"},
    "CH": {"SW"},
    "GB": {"LN", "L"},
    "JP": {"JT", "T"},
    "CA": {"CT", "TO"},
    "KR": {"KS"},
}

_EXCHCODE_TO_YF_SUFFIX = {
    "HE": ".HE", "AS": ".AS", "GY": ".DE", "DE": ".DE",
    "DC": ".CO", "CO": ".CO", "SS": ".ST", "ST": ".ST",
    "OL": ".OL", "SW": ".SW", "LN": ".L",  "L":  ".L",
    "CT": ".TO", "TO": ".TO", "JT": ".T",  "T":  ".T",
    "KS": ".KS",
    "UN": "", "UQ": "", "UA": "", "UR": "", "US": "",
}


def _fetch_openfigi_ticker(isin: str, timeout: int = 4) -> str | None:
    """Query OpenFIGI for a Yahoo Finance ticker suggestion for the given ISIN."""
    try:
        body = json.dumps([{"idType": "ID_ISIN", "idValue": isin}]).encode()
        req = urllib.request.Request(
            "https://api.openfigi.com/v3/mapping",
            data=body,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            result = json.loads(resp.read())
        records = result[0].get("data", [])
        if not records:
            return None
        stocks = [r for r in records if r.get("securityType2") == "Common Stock"] or records
        isin_prefix = isin[:2]
        preferred = _ISIN_PREFIX_TO_EXCHCODES.get(isin_prefix, set())
        match = next((r for r in stocks if r.get("exchCode") in preferred), stocks[0])
        ticker = match.get("ticker", "")
        suffix = _EXCHCODE_TO_YF_SUFFIX.get(match.get("exchCode", ""), "")
        return (ticker + suffix) if ticker else None
    except Exception:
        return None


# Preferred Yahoo Finance exchange codes for name-based ticker selection.
# Primary sort key is still the search score; this list breaks ties in favour of
# liquid, EUR-friendly exchanges over peripheral or non-EUR venues (e.g. Swiss CHF).
_YF_EXCHANGE_PREFERENCE = [
    "NMS", "NYQ", "NCM", "NGM",   # US — NASDAQ / NYSE variants
    "HEL",                          # Helsinki
    "GER", "FRA",                   # Frankfurt / XETRA
    "AMS",                          # Amsterdam
    "MIL",                          # Milan
    "PAR",                          # Paris
    "STO",                          # Stockholm
    "CPH",                          # Copenhagen
    "OSL",                          # Oslo
    "ISE",                          # Ireland
]


def _search_yf_ticker(name: str) -> tuple[str | None, str | None]:
    """Search Yahoo Finance by instrument name.

    Returns (symbol, exchange_label) for the best-scoring equity/ETF/fund result,
    or (None, None) if nothing was found.  The symbol already contains the correct
    Yahoo Finance exchange suffix (e.g. 'SIILI.HE', 'NVDA'), so no further mapping
    is needed.

    When multiple results share the same score, the entry whose exchange appears
    earliest in _YF_EXCHANGE_PREFERENCE wins, avoiding peripheral (e.g. CHF-listed)
    cross-listings for EUR-based portfolios.
    """
    def _exchange_rank(r: dict) -> int:
        try:
            return _YF_EXCHANGE_PREFERENCE.index(r.get("exchange", ""))
        except ValueError:
            return len(_YF_EXCHANGE_PREFERENCE)   # unlisted exchanges go last

    try:
        results = yf.Search(name, max_results=8).quotes
        if not results:
            return None, None
        candidates = [
            r for r in results
            if r.get("quoteType") in ("EQUITY", "ETF", "MUTUALFUND")
        ] or results
        best = min(candidates, key=lambda r: (-r.get("score", 0), _exchange_rank(r)))
        label = best.get("exchDisp") or best.get("exchange") or ""
        return best.get("symbol"), label
    except Exception:
        return None, None


def resolve_tickers(positions, isin_map):
    updated = dict(isin_map)

    def _map_key(p):
        return p["isin"] if p["isin"] else f"NAME:{p['instrument']}"

    unmapped = [(p["instrument"], _map_key(p)) for p in positions if _map_key(p) not in updated]
    if not unmapped:
        if not Path(ISIN_TICKER_MAP_FILE).exists():
            save_isin_map(updated)
        return updated

    # ── Name-based positions (no ISIN): auto-resolve via Yahoo Finance search ──
    name_based = [(n, k) for n, k in unmapped if k.startswith("NAME:")]
    isin_based  = [(n, k) for n, k in unmapped if not k.startswith("NAME:")]

    if name_based:
        print(f"Auto-resolving {len(name_based)} name-based positions via Yahoo Finance...")
        unresolved = []
        for name, map_key in name_based:
            symbol, exch = _search_yf_ticker(name)
            if symbol:
                updated[map_key] = symbol
                print(f"   {name:<48} → {symbol:<12} ({exch})")
            else:
                unresolved.append((name, map_key))
        print()

        if unresolved:
            print("Could not auto-resolve — enter ticker or '-' to skip:")
            for name, map_key in unresolved:
                user_input = input(f"  {name} -> ").strip()
                if user_input and user_input != "-":
                    updated[map_key] = user_input
            print()

    # ── ISIN-based positions: interactive flow with OpenFIGI suggestions ──
    if isin_based:
        print("--- Unmapped ISINs ---")
        print("Suggestions from OpenFIGI shown where available.")
        print("Press Enter to accept, type to override, or '-' to skip.\n")
        for name, map_key in isin_based:
            suggestion = _fetch_openfigi_ticker(map_key)
            if suggestion:
                prompt = f"  {name} ({map_key})\n  Suggestion: {suggestion}\n  Accept [Enter], override, or '-' to skip: "
            else:
                prompt = f"  {name} ({map_key}) -> "
            user_input = input(prompt).strip()
            if suggestion:
                ticker = None if user_input == "-" else (user_input or suggestion)
            else:
                ticker = user_input or None
            if ticker:
                updated[map_key] = ticker
        print()

    save_isin_map(updated)
    print(f"   Saved to {ISIN_TICKER_MAP_FILE}\n")
    return updated


def enrich_positions(positions, isin_map):
    print("Fetching data from Yahoo Finance...")
    enriched = []
    for pos in positions:
        map_key = pos["isin"] if pos["isin"] else f"NAME:{pos['instrument']}"
        ticker = isin_map.get(map_key)
        if not ticker:
            continue
        try:
            info = yf.Ticker(ticker).info
            enriched.append({
                **pos, "ticker": ticker,
                "sector": info.get("sector", "Unknown"),
                "country": info.get("country", "Unknown"),
                "industry": info.get("industry", "Unknown"),
                "full_name": info.get("longName") or info.get("shortName", pos["instrument"]),
            })
        except Exception as e:
            print(f"   Warning: Failed {ticker}: {e}")
            enriched.append({
                **pos, "ticker": ticker,
                "sector": "Unknown", "country": "Unknown",
                "industry": "Unknown", "full_name": pos["instrument"],
            })
    print(f"   Enriched {len(enriched)} positions\n")
    return enriched


def analyze_allocation(enriched, cash_eur):
    total_equity = sum(p["market_value_eur"] for p in enriched)
    total_portfolio = total_equity + cash_eur
    sector_alloc = defaultdict(float)
    geo_alloc = defaultdict(float)
    for p in enriched:
        sector_alloc[p["sector"]] += p["market_value_eur"]
        geo_alloc[p["country"]] += p["market_value_eur"]
    return {
        "total_portfolio_eur": total_portfolio, "total_equity_eur": total_equity,
        "cash_eur": cash_eur,
        "cash_pct": cash_eur / total_portfolio * 100 if total_portfolio else 0,
        "sector": {k: {"value_eur": v, "pct": v / total_equity * 100}
                   for k, v in sorted(sector_alloc.items(), key=lambda x: -x[1])},
        "geography": {k: {"value_eur": v, "pct": v / total_equity * 100}
                      for k, v in sorted(geo_alloc.items(), key=lambda x: -x[1])},
    }


def calculate_xirr(cash_flows_file, current_value):
    if not Path(cash_flows_file).exists():
        return None
    try:
        from scipy.optimize import brentq
    except ImportError:
        return None
    with open(cash_flows_file) as f:
        flows = json.load(f)
    if not flows:
        return None
    entries = [(datetime.strptime(cf["date"], "%Y-%m-%d").date(), cf["amount"]) for cf in flows]
    entries.append((date.today(), -current_value))
    d0 = entries[0][0]
    years = [(d - d0).days / 365.25 for d, _ in entries]
    amounts = [a for _, a in entries]
    def npv(r):
        return sum(a / (1 + r) ** t for a, t in zip(amounts, years))
    try:
        return brentq(npv, -0.99, 10.0)
    except Exception:
        return None


def compute_correlation_matrix(enriched, start_date, end_date):
    tickers = [p["ticker"] for p in enriched]
    if len(tickers) < 2:
        return None
    print(f"Fetching price history {start_date} \u2192 {end_date}...")
    try:
        prices = yf.download(
            tickers, start=start_date.isoformat(), end=end_date.isoformat(),
            auto_adjust=True, progress=False
        )["Close"]
    except Exception as e:
        print(f"   Failed to fetch price history: {e}")
        return None
    if prices.empty:
        return None
    if isinstance(prices, pd.Series):
        prices = prices.to_frame(name=tickers[0])
    valid_cols = [c for c in prices.columns if prices[c].dropna().shape[0] >= 60]
    dropped = set(prices.columns) - set(valid_cols)
    if dropped:
        print(f"   Insufficient data for: {', '.join(str(d) for d in dropped)}")
    prices = prices[valid_cols]
    if prices.shape[1] < 2:
        return None
    # Compute returns without row-wise dropna so tickers from different exchange
    # calendars don't wipe each other's data.  Pairwise min_periods=60 ensures
    # each (ticker, ticker) pair uses only the days when both actually traded.
    returns     = np.log(prices / prices.shift(1))        # NaN rows kept
    corr_matrix = returns.corr(min_periods=60)            # pairwise, NaN-safe
    ann_vol     = returns.std() * np.sqrt(252)            # per-column, NaN-safe

    # Portfolio risk math (w @ Cov @ w) requires a fully populated covariance matrix.
    # Drop any ticker whose covariance column still contains NaN after pairwise
    # computation — this means it shares fewer than 60 trading days with at least
    # one peer (e.g. a non-EUR exchange with a very different market calendar).
    cov_full     = returns.cov(min_periods=60) * 252
    risk_tickers = [c for c in cov_full.columns if cov_full[c].notna().all()]
    excluded     = set(cov_full.columns) - set(risk_tickers)
    if excluded:
        print(f"   Excluded from risk contribution (insufficient data overlap): "
              f"{', '.join(str(e) for e in sorted(excluded))}")
    cov_matrix = cov_full.loc[risk_tickers, risk_tickers]

    equity_total = sum(p["market_value_eur"] for p in enriched)
    ticker_to_weight = {p["ticker"]: p["market_value_eur"] / equity_total
                        for p in enriched if p["ticker"] in risk_tickers}
    ordered_tickers = list(cov_matrix.columns)
    w = np.array([ticker_to_weight.get(t, 0) for t in ordered_tickers])
    w = w / w.sum()
    port_var = w @ cov_matrix.values @ w
    port_vol = np.sqrt(port_var)
    mcr = (cov_matrix.values @ w) / port_vol
    ctr = w * mcr
    pct_ctr = ctr / port_vol * 100
    trading_days = int(returns.count().median())   # typical per-ticker non-NaN count
    print(f"   Computed from {trading_days} trading days")
    print(f"   Annualized portfolio volatility: {port_vol * 100:.1f}%\n")
    return {
        "corr_matrix": corr_matrix, "cov_matrix": cov_matrix, "ann_vol": ann_vol,
        "returns": returns, "tickers": ordered_tickers, "weights": w,
        "port_vol": port_vol, "pct_ctr": pct_ctr, "trading_days": trading_days,
        "start_date": start_date, "end_date": end_date,
    }


def print_report(enriched, allocation, xirr):
    print("=" * 70)
    print("  PORTFOLIO ANALYSIS REPORT")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 70)
    print(f"\n  Total portfolio:  EUR {allocation['total_portfolio_eur']:>12,.2f}")
    print(f"  Equity positions: EUR {allocation['total_equity_eur']:>12,.2f}")
    print(f"  Cash:             EUR {allocation['cash_eur']:>12,.2f}  ({allocation['cash_pct']:.1f}%)")
    if xirr is not None:
        print(f"  XIRR:              {xirr * 100:>11.2f}%")
    equity_total = allocation["total_equity_eur"]
    print(f"\n  {'Ticker':<10} {'Name':<25} {'Value EUR':>10} {'Weight':>8} {'P/L EUR':>10}")
    print(f"  {'-'*10} {'-'*25} {'-'*10} {'-'*8} {'-'*10}")
    for p in sorted(enriched, key=lambda x: -x["market_value_eur"]):
        weight = p["market_value_eur"] / equity_total * 100 if equity_total else 0
        name = p["full_name"][:25]
        print(f"  {p['ticker']:<10} {name:<25} {p['market_value_eur']:>10,.0f} {weight:>7.1f}% {p['pnl_eur']:>+10,.0f}")
    print(f"\n  SECTOR ALLOCATION")
    for sector, data in allocation["sector"].items():
        bar = "#" * int(data["pct"] / 2)
        print(f"  {sector:<28} {data['pct']:>5.1f}%  {bar}")
    print(f"\n  GEOGRAPHY ALLOCATION")
    for country, data in allocation["geography"].items():
        bar = "#" * int(data["pct"] / 2)
        print(f"  {country:<28} {data['pct']:>5.1f}%  {bar}")
    print(f"\n{'=' * 70}\n")


def print_correlation_report(corr_data):
    corr = corr_data["corr_matrix"]
    vol = corr_data["ann_vol"]
    tickers = corr_data["tickers"]
    w = corr_data["weights"]
    pct_ctr = corr_data["pct_ctr"]
    start_date = corr_data["start_date"]
    end_date = corr_data["end_date"]
    print(f"  CORRELATION MATRIX ({start_date} \u2192 {end_date}, daily log returns)")
    print("-" * 70)
    header = "            " + "".join(f"{t:>10}" for t in tickers)
    print(header)
    for i, t in enumerate(tickers):
        row = f"  {t:<10}" + "".join(f"{corr.iloc[i, j]:>10.2f}" for j in range(len(tickers)))
        print(row)
    print(f"\n  RISK DECOMPOSITION")
    print(f"  {'Ticker':<10} {'Weight':>8} {'Ann.Vol':>10} {'Risk Contr.':>12}")
    print(f"  {'-'*10} {'-'*8} {'-'*10} {'-'*12}")
    for i, t in enumerate(tickers):
        print(f"  {t:<10} {w[i]*100:>7.1f}% {vol.iloc[i]*100:>9.1f}% {pct_ctr[i]:>11.1f}%")
    print(f"\n  Portfolio volatility: {corr_data['port_vol'] * 100:.1f}%")
    print(f"  Based on {corr_data['trading_days']} trading days\n")


def generate_charts(enriched, allocation):
    fig = plt.figure(figsize=(16, 12), facecolor="white")
    fig.suptitle(f"Portfolio Analysis  -  {datetime.now().strftime('%Y-%m-%d')}",
                 fontsize=16, fontweight="bold", y=0.97)
    gs = fig.add_gridspec(2, 2, hspace=0.40, wspace=0.30,
                          left=0.07, right=0.95, top=0.90, bottom=0.08)
    ax1 = fig.add_subplot(gs[0, :])
    sorted_pos = sorted(enriched, key=lambda x: x["market_value_eur"])
    equity_total = allocation["total_equity_eur"]
    labels = [p["ticker"] for p in sorted_pos]
    values = [p["market_value_eur"] for p in sorted_pos]
    weights = [v / equity_total * 100 for v in values]
    pnl = [p["pnl_eur"] for p in sorted_pos]
    bar_colors = ["#16a34a" if pl >= 0 else "#dc2626" for pl in pnl]
    bars = ax1.barh(labels, values, color=bar_colors, edgecolor="white", height=0.6)
    for bar, w, v in zip(bars, weights, values):
        ax1.text(bar.get_width() + equity_total * 0.008, bar.get_y() + bar.get_height() / 2,
                 f"EUR {v:,.0f}  ({w:.1f}%)", va="center", fontsize=9, color="#374151")
    ax1.set_title("Holdings by Market Value", fontsize=12, fontweight="bold", loc="left", pad=10)
    ax1.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"EUR {x:,.0f}"))
    ax1.spines[["top", "right"]].set_visible(False)
    ax1.set_xlim(0, max(values) * 1.30)

    ax1.legend(handles=[Patch(color="#16a34a", label="P/L >= 0"), Patch(color="#dc2626", label="P/L < 0")],
               loc="lower right", fontsize=8)

    ax2 = fig.add_subplot(gs[1, 0])
    s_labels = list(allocation["sector"].keys())
    s_values = [allocation["sector"][s]["value_eur"] for s in s_labels]
    s_pcts = [allocation["sector"][s]["pct"] for s in s_labels]
    wedges, _ = ax2.pie(s_values, colors=COLORS[:len(s_labels)], startangle=90,
                        wedgeprops=dict(width=0.45, edgecolor="white", linewidth=2))
    ax2.set_title("Sector Allocation", fontsize=12, fontweight="bold", pad=12)
    ax2.legend(wedges, [f"{s}  ({p:.1f}%)" for s, p in zip(s_labels, s_pcts)],
               loc="center", fontsize=8, frameon=False)

    ax3 = fig.add_subplot(gs[1, 1])
    g_labels = list(allocation["geography"].keys())
    g_values = [allocation["geography"][g]["value_eur"] for g in g_labels]
    g_pcts = [allocation["geography"][g]["pct"] for g in g_labels]
    wedges2, _ = ax3.pie(g_values, colors=COLORS[:len(g_labels)], startangle=90,
                         wedgeprops=dict(width=0.45, edgecolor="white", linewidth=2))
    ax3.set_title("Geography Allocation", fontsize=12, fontweight="bold", pad=12)
    ax3.legend(wedges2, [f"{g}  ({p:.1f}%)" for g, p in zip(g_labels, g_pcts)],
               loc="center", fontsize=8, frameon=False)

    total = allocation["total_portfolio_eur"]
    fig.text(0.5, 0.935,
             f"Total: EUR {total:,.2f}   |   Equity: EUR {equity_total:,.0f}   |   Cash: EUR {allocation['cash_eur']:,.0f} ({allocation['cash_pct']:.1f}%)",
             ha="center", fontsize=10, color="#6b7280")
    plt.savefig(CHARTS_FILE, dpi=150, bbox_inches="tight")
    print(f"   Charts saved to {CHARTS_FILE}")


def generate_correlation_charts(corr_data):
    corr = corr_data["corr_matrix"]
    tickers = corr_data["tickers"]
    w = corr_data["weights"]
    pct_ctr = corr_data["pct_ctr"]
    vol = corr_data["ann_vol"]
    n = len(tickers)
    fig, axes = plt.subplots(1, 2, figsize=(16, 7), facecolor="white",
                              gridspec_kw={"width_ratios": [1.2, 1], "wspace": 0.35})
    date_label = f"{corr_data['start_date']} \u2192 {corr_data['end_date']}"
    fig.suptitle(f"Correlation & Risk  \u2014  {date_label}  ({corr_data['trading_days']} days)",
                 fontsize=14, fontweight="bold", y=0.97)
    ax1 = axes[0]
    im = ax1.imshow(corr.values, cmap=plt.cm.RdYlGn_r, vmin=-1, vmax=1, aspect="equal")
    ax1.set_xticks(range(n))
    ax1.set_yticks(range(n))
    ax1.set_xticklabels(tickers, fontsize=9, rotation=45, ha="right")
    ax1.set_yticklabels(tickers, fontsize=9)
    for i in range(n):
        for j in range(n):
            val = corr.iloc[i, j]
            color = "white" if abs(val) > 0.6 else "black"
            ax1.text(j, i, f"{val:.2f}", ha="center", va="center",
                     fontsize=9, fontweight="bold" if i == j else "normal", color=color)
    fig.colorbar(im, ax=ax1, fraction=0.046, pad=0.04).set_label("Correlation", fontsize=9)
    ax1.set_title("Pairwise Correlation", fontsize=12, fontweight="bold", pad=12)

    ax2 = axes[1]
    x = np.arange(n)
    bw = 0.35
    ax2.bar(x - bw / 2, w * 100, bw, label="Portfolio Weight", color="#2563eb", alpha=0.8)
    risk_bars = ax2.bar(x + bw / 2, pct_ctr, bw, label="Risk Contribution", color="#dc2626", alpha=0.8)
    ax2.set_xticks(x)
    ax2.set_xticklabels(tickers, fontsize=9, rotation=45, ha="right")
    ax2.set_ylabel("Percentage (%)", fontsize=9)
    ax2.set_title("Weight vs Risk Contribution", fontsize=12, fontweight="bold", pad=12)
    ax2.legend(fontsize=9, loc="upper right")
    ax2.spines[["top", "right"]].set_visible(False)
    for bar, v in zip(risk_bars, vol):
        ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.8,
                 f"vol={v*100:.0f}%", ha="center", fontsize=7, color="#6b7280")
    plt.savefig(CORRELATION_CHARTS_FILE, dpi=150, bbox_inches="tight")
    print(f"   Correlation charts saved to {CORRELATION_CHARTS_FILE}")


def compute_expanded_corr(candidates, corr_data, start_date, end_date):
    """Download candidate price history and build a combined correlation dataset."""
    print(f"Fetching price history for candidates: {', '.join(candidates)}...")
    try:
        raw = yf.download(
            candidates, start=start_date.isoformat(), end=end_date.isoformat(),
            auto_adjust=True, progress=False,
        )["Close"]
        if isinstance(raw, pd.Series):
            raw = raw.to_frame(name=candidates[0])
    except Exception as e:
        print(f"   Failed to fetch candidate prices: {e}")
        return None

    if raw.empty:
        return None

    valid_candidates = [c for c in raw.columns if raw[c].dropna().shape[0] >= 60]
    dropped = set(raw.columns) - set(valid_candidates)
    if dropped:
        print(f"   Insufficient data for: {', '.join(str(d) for d in dropped)}")

    already_held = [c for c in valid_candidates if c in corr_data["returns"].columns]
    if already_held:
        print(f"   Already in portfolio, skipping: {', '.join(already_held)}")
    valid_candidates = [c for c in valid_candidates if c not in corr_data["returns"].columns]

    if not valid_candidates:
        return None

    cand_returns = np.log(raw[valid_candidates] / raw[valid_candidates].shift(1)).dropna()
    combined = corr_data["returns"].join(cand_returns, how="inner")
    if combined.shape[0] < 60:
        print("   Not enough overlapping trading days for expanded matrix.")
        return None

    print(f"   Expanded matrix: {len(combined.columns)} stocks, {combined.shape[0]} trading days\n")
    return {
        "corr_matrix": combined.corr(),
        "ann_vol": combined.std() * np.sqrt(252),
        "returns": combined,
        "existing_tickers": corr_data["tickers"],
        "candidate_tickers": valid_candidates,
        "all_tickers": list(combined.columns),
        "trading_days": combined.shape[0],
        "weights": corr_data["weights"],
        "port_vol": corr_data["port_vol"],
        "start_date": start_date,
        "end_date": end_date,
    }


def generate_expanded_correlation_chart(expanded_data):
    """Plot the full correlation matrix (existing + candidates) with candidates highlighted."""
    corr = expanded_data["corr_matrix"]
    all_tickers = expanded_data["all_tickers"]
    candidate_tickers = expanded_data["candidate_tickers"]
    ann_vol = expanded_data["ann_vol"]
    n = len(all_tickers)

    fig, axes = plt.subplots(1, 2, figsize=(max(16, n * 1.3), 7), facecolor="white",
                             gridspec_kw={"width_ratios": [1.4, 1], "wspace": 0.40})
    date_label = f"{expanded_data['start_date']} \u2192 {expanded_data['end_date']}"
    fig.suptitle(
        f"Expanded Correlation  \u2014  existing + candidates  \u2014  {date_label}  ({expanded_data['trading_days']} days)",
        fontsize=14, fontweight="bold", y=0.97,
    )

    # Left: full pairwise correlation heatmap
    ax1 = axes[0]
    im = ax1.imshow(corr.values, cmap=plt.cm.RdYlGn_r, vmin=-1, vmax=1, aspect="equal")
    ax1.set_xticks(range(n))
    ax1.set_yticks(range(n))
    x_labels = ax1.set_xticklabels(all_tickers, fontsize=9, rotation=45, ha="right")
    y_labels = ax1.set_yticklabels(all_tickers, fontsize=9)
    for lbl in list(x_labels) + list(y_labels):
        if lbl.get_text() in candidate_tickers:
            lbl.set_color("#ea580c")
            lbl.set_fontweight("bold")
    for i in range(n):
        for j in range(n):
            val = corr.iloc[i, j]
            txt_color = "white" if abs(val) > 0.6 else "black"
            ax1.text(j, i, f"{val:.2f}", ha="center", va="center",
                     fontsize=8, fontweight="bold" if i == j else "normal", color=txt_color)
    fig.colorbar(im, ax=ax1, fraction=0.046, pad=0.04).set_label("Correlation", fontsize=9)
    ax1.set_title("Pairwise Correlation  (orange = candidates)",
                  fontsize=11, fontweight="bold", pad=12)

    # Right: individual annualised volatility for all stocks
    ax2 = axes[1]
    vols = ann_vol.reindex(all_tickers) * 100
    bar_colors = ["#ea580c" if t in candidate_tickers else "#2563eb" for t in all_tickers]
    ax2.bar(range(n), vols, color=bar_colors, alpha=0.85, edgecolor="white")
    ax2.set_xticks(range(n))
    ax2.set_xticklabels(all_tickers, fontsize=9, rotation=45, ha="right")
    ax2.set_ylabel("Annualised Volatility (%)", fontsize=9)
    ax2.set_title("Individual Volatility", fontsize=11, fontweight="bold", pad=12)
    ax2.spines[["top", "right"]].set_visible(False)

    ax2.legend(handles=[Patch(color="#2563eb", label="Existing holding"),
                        Patch(color="#ea580c", label="Candidate")],
               fontsize=8, loc="upper right")

    plt.savefig(EXPANDED_CORRELATION_FILE, dpi=150, bbox_inches="tight")
    print(f"   Expanded correlation chart saved to {EXPANDED_CORRELATION_FILE}")


def pick_stocks(candidates, corr_data, enriched, test_weight=0.05, start_date=None, end_date=None):
    """Compare candidate tickers against existing portfolio for diversification."""
    expanded_data = compute_expanded_corr(candidates, corr_data, start_date, end_date)
    if not expanded_data:
        print("No valid candidates to compare.")
        return None

    existing_tickers = corr_data["tickers"]
    combined = expanded_data["returns"]
    results = []

    for ticker in expanded_data["candidate_tickers"]:
        valid_existing = [t for t in existing_tickers if t in combined.columns]
        pair_corrs = {t: combined[ticker].corr(combined[t]) for t in valid_existing}
        avg_corr = float(np.mean(list(pair_corrs.values())))
        max_corr_ticker = max(pair_corrs, key=pair_corrs.get)
        max_corr = float(pair_corrs[max_corr_ticker])

        sim_cov = combined[valid_existing + [ticker]].cov() * 252
        total_w = sum(corr_data["weights"][existing_tickers.index(t)] for t in valid_existing)
        sim_w = np.array(
            [corr_data["weights"][existing_tickers.index(t)] / total_w * (1 - test_weight)
             for t in valid_existing]
            + [test_weight]
        )
        sim_vol = float(np.sqrt(sim_w @ sim_cov.values @ sim_w))

        try:
            info = yf.Ticker(ticker).info
            name = info.get("longName") or info.get("shortName", ticker)
            sector = info.get("sector", "Unknown")
        except Exception:
            name, sector = ticker, "Unknown"

        results.append({
            "ticker": ticker, "name": name, "sector": sector,
            "avg_corr": avg_corr, "max_corr": max_corr, "max_corr_ticker": max_corr_ticker,
            "pair_corrs": pair_corrs, "sim_vol": sim_vol,
            "vol_delta": (sim_vol - corr_data["port_vol"]) * 100,
        })

    if not results:
        print("No valid candidates to compare.")
        return expanded_data

    results.sort(key=lambda x: x["max_corr"])
    print_stock_pick_report(results, corr_data, test_weight)
    return expanded_data


def print_stock_pick_report(results, corr_data, test_weight):
    print("=" * 70)
    print("  STOCK PICKER  —  Diversification Analysis")
    print(f"  Simulated at {test_weight*100:.0f}% allocation "
          f"(existing holdings scaled to {(1-test_weight)*100:.0f}%)")
    print("=" * 70)
    print(f"  Current portfolio volatility: {corr_data['port_vol']*100:.1f}%\n")
    print(f"  {'Rank':<5} {'Ticker':<10} {'Sector':<22} {'Avg Corr':>9} {'Max Corr':>9} {'Sim Vol':>8} {'Vol Δ':>8}  Verdict")
    print(f"  {'-'*5} {'-'*10} {'-'*22} {'-'*9} {'-'*9} {'-'*8} {'-'*8}  {'-'*23}")

    for rank, r in enumerate(results, 1):
        if r["max_corr"] < 0.5:
            verdict = "GOOD FIT"
        elif r["max_corr"] < 0.75:
            verdict = f"NEUTRAL vs {r['max_corr_ticker']}"
        else:
            verdict = f"HIGH OVERLAP vs {r['max_corr_ticker']}"
        delta_str = f"{r['vol_delta']:+.1f}%"
        print(f"  {rank:<5} {r['ticker']:<10} {r['sector'][:22]:<22} "
              f"{r['avg_corr']:>9.2f} {r['max_corr']:>9.2f} {r['sim_vol']*100:>7.1f}% {delta_str:>8}  {verdict}")

    print()
    for r in results:
        print(f"  {r['ticker']}  ({r['name'][:50]})")
        print(f"  {'Holding':<12} {'Correlation':>12}")
        for t, c in sorted(r["pair_corrs"].items(), key=lambda x: x[1]):
            print(f"  {t:<12} {c:>12.2f}")
        print()
    print("=" * 70 + "\n")


def main():
    parser = argparse.ArgumentParser(
        description="Lightweight Portfolio Analyzer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n"
               "  python portfolio_analyzer.py Portfolio_report.pdf\n"
               "  python portfolio_analyzer.py --pick NVDA TSM ASML\n"
               "  python portfolio_analyzer.py Portfolio_report.pdf --pick NVDA TSM",
    )
    parser.add_argument("report", nargs="?",
                        help="Portfolio report file — PDF (Saxo/Mandatum) or CSV (auto-detected if omitted)")
    parser.add_argument("--pick", nargs="+", metavar="TICKER",
                        help="Candidate ticker(s) to evaluate for diversification")
    parser.add_argument("--test-weight", type=float, default=0.05, metavar="W",
                        help="Simulated allocation for each candidate (default: 0.05)")
    parser.add_argument("--start", metavar="YYYY-MM-DD",
                        help="Start date for price history (default: 1 year before --end)")
    parser.add_argument("--end", metavar="YYYY-MM-DD",
                        help="End date for price history (default: today)")
    args = parser.parse_args()

    if args.end:
        try:
            end_date = date.fromisoformat(args.end)
        except ValueError:
            print(f"Invalid --end date: {args.end!r}. Use YYYY-MM-DD.")
            sys.exit(1)
    else:
        end_date = date.today()

    if args.start:
        try:
            start_date = date.fromisoformat(args.start)
        except ValueError:
            print(f"Invalid --start date: {args.start!r}. Use YYYY-MM-DD.")
            sys.exit(1)
    else:
        start_date = end_date - timedelta(days=365)

    if start_date >= end_date:
        print(f"--start ({start_date}) must be before --end ({end_date}).")
        sys.exit(1)

    if args.report:
        input_path = args.report
    else:
        candidates = list(Path(".").glob("Portfolio_*.pdf")) + list(Path(".").glob("Portfolio_*.csv"))
        if candidates:
            input_path = str(candidates[0])
            print(f"Auto-detected: {input_path}\n")
        else:
            parser.print_help()
            sys.exit(1)

    if not Path(input_path).exists():
        print(f"File not found: {input_path}")
        sys.exit(1)

    holdings = load_portfolio(input_path)
    isin_map = load_isin_map()
    isin_map = resolve_tickers(holdings["positions"], isin_map)
    enriched = enrich_positions(holdings["positions"], isin_map)

    if not enriched:
        print("No enriched positions. Check isin_ticker_map.json.")
        sys.exit(1)

    allocation = analyze_allocation(enriched, holdings["cash_eur"])
    xirr = calculate_xirr(CASH_FLOWS_FILE, holdings["account_value_eur"])

    corr_data = compute_correlation_matrix(enriched, start_date, end_date)

    print_report(enriched, allocation, xirr)
    if corr_data:
        print_correlation_report(corr_data)

    generate_charts(enriched, allocation)
    if corr_data:
        generate_correlation_charts(corr_data)

    if args.pick:
        if not corr_data:
            print("Cannot run stock picker: correlation matrix unavailable "
                  "(need >= 2 positions with sufficient history).")
            sys.exit(1)
        expanded_data = pick_stocks(args.pick, corr_data, enriched, test_weight=args.test_weight,
                                    start_date=start_date, end_date=end_date)
        if expanded_data:
            generate_expanded_correlation_chart(expanded_data)

    plt.show()


if __name__ == "__main__":
    main()
